# pip3 install openpyxl

from calendar import c
from http.client import NO_CONTENT
from operator import indexOf
from pickle import NONE
from openpyxl import Workbook, load_workbook
from tomlkit import datetime

# the study stats (both MesCoBraD and CBTI)
wb_ss = load_workbook('study_stats.xlsx')
# accessing values from 'Data type' sheet
ws_ss = wb_ss['Data type']
# accessing values from 'Sheet2' sheet to analyze
wb_ss_s2 = wb_ss['Sheet2']
# accessing the values from sheet 'CBTI-STUDY SUMMARY'
ws3 = wb_ss['CBTI-STUDY SUMMARY']
# accessing the values from sheet 'CBTI-INTERVENTIONS'
ws4 = wb_ss['CBTI-INTERVENTIONS']

# print(wb_ss.sheetnames)

# NIA demographics stats
wb_nia = load_workbook('NIA Demographics.xlsx')
ws_nia = wb_nia.active

# new workbook for displaying statistics for NIA and study patients
wb = Workbook()
ws = wb.active
ws.title = 'Statistics'
ws.append(['NIA Patients MRN and Count', 'MesCoBraD Enrolled',
          'CBTI Enrolled', 'MesCoBraD and CBTI Enrolled', 'NIA- Not in study', ])

# a sheet for tracking appointment status
ws_apts = wb.create_sheet('Appointments')
ws_apts.append(['NPT', 'PSG', 'Actigraphy', '3moNPT', '1YNPT', '1YPSG', 'Interventions Session 1', 'Interventions Session 2', 'Interventions Session 3',
               'Interventions Session 4', 'Interventions Session 5', 'Interventions Session 6', 'Interventions Completion Status'])


# a sheet for tracking missing appointments
ws_missed_appointments = wb.create_sheet('Missing Appointments')
ws_missed_appointments.append(['NPT', 'PSG', 'Actigraphy', '3moNPT', '1YNPT', '1YPSG', 'Interventions Session 1', 'Interventions Session 2', 'Interventions Session 3',
                               'Interventions Session 4', 'Interventions Session 5', 'Interventions Session 6', 'Interventions Completion Status'])
# dictionary for keeping track of whhat has been completed for each patient care/research pathway

enrolled = {'MesCoBraD Enrolled': {'MRN': []}, 'CBTI Enrolled': {
    'MRN': []}, 'MesCoBraD and CBTI Enrolled': {'MRN': []}, 'NIA_total': {'MRN': []}, 'NIA- Not in study': {'MRN': []}}


mesbrad_completed_appointments = {}
cbti_completed_appointments = {}

completed_npt = []
completed_psg = []
completed_actigraphy = []
completed_3mo_npt = []
completed_1y_psg = []
completed_1y_npt = []
completed_cbti_sess1 = []
completed_cbti_sess2 = []
completed_cbti_sess3 = []
completed_cbti_sess4 = []
completed_cbti_sess5 = []
completed_cbti_sess6 = []
completed_cbti_sess_all = []


not_completed_npt = []
not_completed_psg = []
not_completed_actigraphy = []
not_completed_3mo_npt = []
not_completed_1y_psg = []
not_completed_1y_npt = []
not_completed_cbti_sess1 = []
not_completed_cbti_sess2 = []
not_completed_cbti_sess3 = []
not_completed_cbti_sess4 = []
not_completed_cbti_sess5 = []
not_completed_cbti_sess6 = []
not_completed_cbti_sess_all = []


mescobrad_mrn_list = enrolled['MesCoBraD Enrolled'].keys()
cbti_mrn_list = enrolled['CBTI Enrolled'].keys()
nia_mrn_list = enrolled['NIA_total'].keys()
nia_no_study = enrolled['NIA- Not in study'].keys()
nia_both_study = enrolled['MesCoBraD and CBTI Enrolled'].keys()


def NIA_patient_stats():
    ''' # of NIA patients based on demographics sheet'''

    ws['A2'] = 0
    mrn_last_value = 0
    count = 0
    for mrn in ws_nia.rows:
        mrn_value = mrn[0].value
        if mrn_value == 'MRN':
            pass
        else:
            if mrn_value != mrn_last_value:
                count += 1
                ws.append([mrn_value])
                enrolled['NIA_total']['MRN'].append(mrn_value)
                mrn_last_value = mrn_value
    ws['A2'].value = count
    return


def MesCoBraD_enrolled():
    '''MesCoBraD # of enrolled'''
    ws['B2'] = 0
    mrn_last_value = 0
    count = 0
    for mrn in ws_ss.rows:
        if mrn[0].value == 'MRN':
            pass
        else:
            mrn_value = mrn[0].value
            if mrn_value != mrn_last_value:
                count += 1
                mrn_last_value = mrn_value
                enrolled['MesCoBraD Enrolled']['MRN'].append(mrn_value)
                next_open_cell = 'B' + str(count+2)
                ws[next_open_cell] = mrn_value
    ws['B2'].value = count
    return wb.save('statistics.xlsx')


def CBTI_enrolled():
    '''function for CBTI enrolled count'''
    ws['C2'] = 0
    count = 0
    for mrn in ws3.rows:
        mrn_value = mrn[0].value

        enrolled['CBTI Enrolled']['MRN'].append(mrn_value)
        next_open_cell = 'C' + str(count+2)
        ws[next_open_cell] = mrn_value
        count += 1
    ws['C2'] = len(enrolled['CBTI Enrolled']['MRN'])
    return


def pts_CBTI_and_MesCoBrad():
    '''MRN's of patients in both CBTI and MesCoBraD'''

    ws['D2'] = 0
    count = 0
    for mrn in enrolled['NIA_total']['MRN']:
        if mrn in enrolled['CBTI Enrolled']['MRN']:
            if mrn in enrolled['MesCoBraD Enrolled']['MRN']:
                enrolled['MesCoBraD and CBTI Enrolled']['MRN'].append(
                    mrn)
                next_open_cell = 'D' + str(count+2)
                ws[next_open_cell] = mrn
                count += 1
    ws['D2'] = count
    return


def nia_not_in_study():
    '''list of MRN for NIA patients not in any study'''
    nia_no_study = enrolled['NIA- Not in study']['MRN']

    ws['E2'] = 0
    count = 0
    for mrn in enrolled['NIA_total']['MRN']:
        if mrn in enrolled['CBTI Enrolled']['MRN']:
            pass
        elif mrn in enrolled['MesCoBraD Enrolled']['MRN']:
            pass
        else:
            enrolled['NIA- Not in study']['MRN'].append(mrn)
            next_open_cell = 'E' + str(len(nia_no_study)+2)
            ws[next_open_cell] = mrn
            count += 1
    ws['E2'] = count

    return


def finding_column_for_appointments(apt_type):
    column = ""
    if apt_type == 'NPT':
        column = 'A'
    elif apt_type == 'PSG':
        column = 'B'
    elif apt_type == 'Actigraphy':
        column = 'C'
    elif apt_type == '3moNPT':
        column = 'D'
    elif apt_type == '1YNPT':
        column = 'E'
    elif apt_type == '1YPSG':
        column = 'F'
    elif apt_type == 'Interventions Session 1':
        column = 'G'
    elif apt_type == 'Interventions Session 2':
        column = 'H'
    elif apt_type == 'Interventions Session 3':
        column = 'I'
    elif apt_type == 'Interventions Session 4':
        column = 'J'
    elif apt_type == 'Interventions Session 5':
        column = 'K'
    elif apt_type == 'Interventions Session 6':
        column = 'L'
    elif apt_type == 'Interventions Completion Status':
        column = 'M'
    return column


completed_apts_row_dict = {'NPT': 0, 'PSG': 0, 'Actigraphy': 0, '3moNPT': 0, '1YNPT': 0, '1YPSG': 0, 'Interventions Session 1': 0,
                           'Interventions Session 2': 0, 'Interventions Session 3': 0, 'Interventions Session 4': 0, 'Interventions Session 5': 0, 'Interventions Session 6': 0, 'Interventions Completion Status': 0}


def adding_completed_appointments_to_ws(apt_type, mrn, next_open_row):
    '''adding appointment to completed appointments column on worksheet = appointments'''
    column = finding_column_for_appointments(apt_type)
    ws_apts[column + str(next_open_row + 2)] = mrn
    ws_apts[column + '2'] = completed_apts_row_dict[apt_type]

    return


def appointment_lists(apt_type, mrn):
    '''creating lists of completed appointments and adding to appointments worksheet'''
    if apt_type == ' HOME PSG':
        completed_psg.append(mrn)
        apt_type = 'PSG'
    elif apt_type == 'IN-LAB PSG':
        completed_psg.append(mrn)
        apt_type = 'PSG'
    elif apt_type == 'NPT':
        completed_npt.append(mrn)
    elif apt_type == '3moNPT':
        completed_3mo_npt.append(mrn)
    elif apt_type == 'Actigraphy':
        completed_actigraphy.append(mrn)
    elif apt_type == '1YNPT':
        completed_1y_npt.append(mrn)
    elif apt_type == '1YPSG':
        completed_1y_psg.append(mrn)
    elif apt_type == 'Interventions Session 1':
        completed_cbti_sess1.append(mrn)
    elif apt_type == 'Interventions Session 2':
        completed_cbti_sess2.append(mrn)
    elif apt_type == 'Interventions Session 3':
        completed_cbti_sess3.append(mrn)
    elif apt_type == 'Interventions Session 4':
        completed_cbti_sess4.append(mrn)
    elif apt_type == 'Interventions Session 5':
        completed_cbti_sess5.append(mrn)
    elif apt_type == 'Interventions Session 6':
        completed_cbti_sess6.append(mrn)
    else:
        return
    completed_apts_row_dict[apt_type] += 1
    next_open_row = completed_apts_row_dict[apt_type]
    adding_completed_appointments_to_ws(apt_type, mrn, next_open_row)

    return


not_completed_apts_row_dict = {'NPT': 0, 'PSG': 0, 'Actigraphy': 0, '3moNPT': 0, '1YNPT': 0, '1YPSG': 0, 'Interventions Session 1': 0,
                               'Interventions Session 2': 0, 'Interventions Session 3': 0, 'Interventions Session 4': 0, 'Interventions Session 5': 0, 'Interventions Session 6': 0, 'Interventions Completion Status': 0}


def adding_not_completed_appointments_to_ws(apt_type, mrn, next_open_row):
    '''adding missed appointment to missed appointments column on worksheet = missing appointments'''
    column = finding_column_for_appointments(apt_type)
    if mrn == "ID Number" or mrn == None:
        return
    else:
        ws_missed_appointments[column + str(next_open_row + 1)] = mrn
        ws_missed_appointments[column +
                               '2'] = not_completed_apts_row_dict[apt_type] - 1

    return


def incomplete_appointments_list(mrn, apt_type):
    '''creating lists of patients with missing appointments sorted by type'''
    if apt_type == 'ID Number':
        return
    elif apt_type == ' HOME PSG':
        not_completed_psg.append(mrn)
        apt_type = 'PSG'
    elif apt_type == 'IN-LAB PSG':
        not_completed_psg.append(mrn)
        apt_type = 'PSG'
    elif apt_type == 'NPT':
        not_completed_npt.append(mrn)
    elif apt_type == '3moNPT':
        not_completed_3mo_npt.append(mrn)
    elif apt_type == 'Actigraphy':
        not_completed_actigraphy.append(mrn)
    elif apt_type == '1YNPT':
        not_completed_1y_npt.append(mrn)
    elif apt_type == '1YPSG':
        not_completed_1y_psg.append(mrn)
    elif apt_type == 'Interventions Session 1':
        not_completed_cbti_sess1.append(mrn)
    elif apt_type == 'Interventions Session 2':
        not_completed_cbti_sess2.append(mrn)
    elif apt_type == 'Interventions Session 3':
        not_completed_cbti_sess3.append(mrn)
    elif apt_type == 'Interventions Session 4':
        not_completed_cbti_sess4.append(mrn)
    elif apt_type == 'Interventions Session 5':
        not_completed_cbti_sess5.append(mrn)
    elif apt_type == 'Interventions Session 6':
        not_completed_cbti_sess6.append(mrn)
    else:
        return

    not_completed_apts_row_dict[apt_type] += 1
    next_open_row = not_completed_apts_row_dict[apt_type]
    adding_not_completed_appointments_to_ws(
        apt_type, mrn, next_open_row)
    return


def mescobrad_appointments():
    '''required_appointments':'NPT', 'PSG', 'Actigraphy', '1YNPT'; only some require 1YPSG'''

    pt_count = 0
    current_mrn = 0
    count = 0

    for mrn in ws_ss:

        row = count + 2
        c_cell = ws_ss['C' + str(row)].value
        b_cell = list(str(ws_ss['B' + str(row)].value))
        mrn_value = mrn[0].value

        if mrn_value != "MRN":
            count += 1
            if len(b_cell) == 19:
                if mrn_value != current_mrn:
                    mesbrad_completed_appointments[mrn_value] = {
                        c_cell: ws_ss['B' + str(row)].value}
                    appointment_lists(c_cell, mrn_value)
                    current_mrn = mrn_value
                    pt_count += 1
                elif mrn_value == current_mrn:
                    mesbrad_completed_appointments[mrn_value] = {
                        c_cell: ws_ss['B' + str(row)].value}
                    appointment_lists(c_cell, mrn_value)
            else:
                if mrn_value != current_mrn:
                    incomplete_appointments_list(
                        mrn_value, c_cell)
                    pt_count += 1
                    current_mrn = mrn_value
                else:
                    incomplete_appointments_list(
                        mrn_value, c_cell)
    return


def cbti_appointments(worksheet, apt_or_session):
    ''' CBTI appointments'''
    pt_count = 0

    for mrn in worksheet:
        mrn_value = mrn[0].value
        row = pt_count + 2
        cell = []
        if apt_or_session == 'apt':
            columns = {'F': '3moNPT', 'G': '1YNPT', 'H': '1YPSG'}
        elif apt_or_session == 'session':
            columns = {'B': 'Interventions Session 1', 'C': 'Interventions Session 2', 'D': 'Interventions Session 3',
                       'E': 'Interventions Session 4', 'F': 'Interventions Session 5', 'G': 'Interventions Session 6'}
        for column in columns.keys():
            cell = worksheet[column + str(row)].value
            apt_type = columns[column]
            if cell == 1:
                cbti_completed_appointments[mrn_value] = {
                    columns[column]: worksheet[cell]}
                appointment_lists(apt_type, mrn_value)
            else:
                incomplete_appointments_list(
                    mrn_value, apt_type)
        pt_count += 1
    return


NIA_patient_stats()
MesCoBraD_enrolled()
CBTI_enrolled()
pts_CBTI_and_MesCoBrad()
nia_not_in_study()
mescobrad_appointments()
cbti_appointments(ws3, 'apt')
cbti_appointments(ws4, 'session')
wb.save('statistics.xlsx')
